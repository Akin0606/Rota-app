"""Branded PDF and Excel exports of a published rota.

Both formats render the same staff×day matrix and honour the axis orientation
chosen in the app:
  - "staff-rows" (default): staff down the left, days across the top
  - "day-rows"  (inverted): days down the left, staff across the top

Each cell shows the shift name and a compact time range, e.g. "Day 2–6pm".
The PDF is branded (venue + week header, Crewplan wordmark in the corner) so
it can be shared outside the app.
"""

from __future__ import annotations

import io
import re
from datetime import date, timedelta

from fpdf import FPDF

from services import shift_bounds

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DAY_SHORT = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

ACCENT = (255, 77, 0)  # #FF4D00
INK = (17, 24, 39)  # #111827
INK_MUTED = (107, 114, 128)  # #6b7280
HAIRLINE = (229, 231, 235)  # #e5e7eb

_CLOCK_RE = re.compile(r"^(\d{1,2})(?::(\d{2}))?\s*(am|pm)?$", re.IGNORECASE)


def _parse_clock(text: str):
    m = _CLOCK_RE.match((text or "").strip())
    if not m:
        return None
    hour = int(m.group(1))
    minute = int(m.group(2)) if m.group(2) else 0
    suffix = (m.group(3) or "").lower()
    if hour > 23 or minute > 59:
        return None
    return hour, minute, suffix


def _label(hour: int, minute: int, suffix: str) -> str:
    # Normalise to a 12-hour display hour when we have no explicit am/pm.
    disp = hour % 12 or 12 if not suffix else hour
    return f"{disp}:{minute:02d}" if minute else f"{disp}"


def compact_time_range(start: str, end: str) -> str:
    """"2:00pm"/"6:00pm" -> "2–6pm"; "9am"/"5pm" -> "9am–5pm". Falls back to the
    raw values joined with an en dash when either side isn't a clock time."""
    s = _parse_clock(start)
    e = _parse_clock(end)
    if not s or not e:
        raw_start = (start or "").strip()
        raw_end = (end or "").strip()
        return f"{raw_start}–{raw_end}".strip("–")
    sh, sm, ss = s
    eh, em, es = e
    if ss and es and ss == es:
        return f"{_label(sh, sm, ss)}–{_label(eh, em, es)}{ss}"
    return f"{_label(sh, sm, ss)}{ss}–{_label(eh, em, es)}{es}"


def _cell_text(shift: dict | None, day_index: int, shift_days_idx: dict | None) -> str:
    if not shift:
        return ""
    start, end = shift_bounds.bounds_for(shift, day_index, shift_days_idx)
    return f"{shift['name']} {compact_time_range(start, end)}".strip()


# fpdf2's core fonts are latin-1 only, so map the handful of unicode
# punctuation marks we emit down to safe equivalents for the PDF path.
# (Excel and the on-screen grid keep the nicer en dash.)
_L1_MAP = {"–": "-", "—": "-", "…": "...", "’": "'", "“": '"', "”": '"'}


def _pdf_safe(text: str) -> str:
    for k, v in _L1_MAP.items():
        text = text.replace(k, v)
    return text.encode("latin-1", "replace").decode("latin-1")


def _week_range_label(week_start: date) -> str:
    end = week_start + timedelta(days=6)
    if week_start.month == end.month:
        return f"{week_start.day}–{end.day} {end.strftime('%b %Y')}"
    return f"{week_start.strftime('%d %b')} – {end.strftime('%d %b %Y')}"


def _build_matrix(shifts: list[dict], assignments: list[dict], shift_days_idx: dict | None = None):
    """Returns lookup (staff_id, day_index) -> {"shift", "text"}.

    Per-day times are resolved here, where the day_index is known, so the
    orientation-agnostic draw code never needs the day. `text` is the resolved
    cell label; `shift` is kept for the colour fill.
    """
    shifts_by_id = {s["id"]: s for s in shifts}
    cell: dict[tuple[str, int], dict] = {}
    for a in assignments:
        shift = shifts_by_id.get(a.get("shift_id"))
        if shift is not None:
            day = a["day_index"]
            cell[(a["staff_id"], day)] = {
                "shift": shift,
                "text": _cell_text(shift, day, shift_days_idx),
            }
    return cell


# --- PDF -------------------------------------------------------------------


class _RotaPDF(FPDF):
    def __init__(self, venue_name: str, week_label: str):
        super().__init__(orientation="L", unit="mm", format="A4")
        self.venue_name = venue_name
        self.week_label = week_label
        self.set_auto_page_break(auto=True, margin=18)
        self.set_margins(12, 12, 12)

    def header(self):
        self.set_xy(12, 12)
        self.set_text_color(*INK)
        self.set_font("Helvetica", "B", 16)
        self.cell(0, 8, _pdf_safe(self.venue_name), new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(*INK_MUTED)
        self.set_font("Helvetica", "", 11)
        self.cell(0, 6, _pdf_safe(f"Rota for week {self.week_label}"), new_x="LMARGIN", new_y="NEXT")
        self.ln(2)
        y = self.get_y()
        self.set_draw_color(*HAIRLINE)
        self.line(12, y, self.w - 12, y)
        self.ln(3)

    def footer(self):
        # Crewplan wordmark in the bottom-right corner — this doc is shared
        # outside the app, so it carries the brand.
        self.set_y(-14)
        self.set_font("Helvetica", "B", 11)
        wordmark = "crewplan"
        dot = "."
        w_word = self.get_string_width(wordmark)
        w_dot = self.get_string_width(dot)
        x = self.w - 12 - (w_word + w_dot)
        self.set_x(x)
        self.set_text_color(*INK)
        self.cell(w_word, 6, wordmark, new_x="RIGHT", new_y="TOP")
        self.set_text_color(*ACCENT)
        self.cell(w_dot, 6, dot)


def _name_with_tag(person: dict) -> str:
    return f"{person['name']} (U18)" if person.get("is_under_18") else person["name"]


def build_rota_pdf(
    *,
    venue_name: str,
    week_start: date,
    shifts: list[dict],
    staff: list[dict],
    assignments: list[dict],
    leave: dict[str, set[int]] | None = None,
    orientation: str = "staff-rows",
    shift_days: dict | None = None,
) -> bytes:
    leave = leave or {}
    week_label = _week_range_label(week_start)
    cell = _build_matrix(shifts, assignments, shift_days)

    pdf = _RotaPDF(venue_name, week_label)
    pdf.add_page()

    if orientation == "day-rows":
        row_headers = [DAY_SHORT[i] for i in range(7)]
        col_headers = [_name_with_tag(s) for s in staff]
        col_keys = [s["id"] for s in staff]

        def lookup(r: int, c: int):
            return cell.get((col_keys[c], r))

        def is_leave(r: int, c: int) -> bool:
            return r in leave.get(col_keys[c], ())

        _draw_table(pdf, row_headers, col_headers, lookup, is_leave)
    else:
        # Grouped by role, same as the on-screen matrix — a manager scanning a
        # printed rota reads it the same way they read the app.
        groups: dict[str, list[dict]] = {}
        for s in staff:
            groups.setdefault(s.get("role") or "Staff", []).append(s)
        col_headers = [DAY_SHORT[i] for i in range(7)]
        for role, members in groups.items():
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*INK_MUTED)
            pdf.cell(0, 7, _pdf_safe(role), new_x="LMARGIN", new_y="NEXT")
            row_headers = [_name_with_tag(m) for m in members]
            row_keys = [m["id"] for m in members]

            def lookup(r: int, c: int, _keys=row_keys):
                return cell.get((_keys[r], c))

            def is_leave(r: int, c: int, _keys=row_keys) -> bool:
                return c in leave.get(_keys[r], ())

            _draw_table(pdf, row_headers, col_headers, lookup, is_leave)
            pdf.ln(2)

    if not assignments:
        pdf.ln(4)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*INK_MUTED)
        pdf.cell(0, 6, "No shifts assigned for this week yet.")

    out = pdf.output()
    return bytes(out)


def _draw_table(pdf: _RotaPDF, row_headers, col_headers, lookup, is_leave=None) -> None:
    is_leave = is_leave or (lambda r, c: False)
    n_cols = len(col_headers)
    usable = pdf.w - 24  # left+right margins
    label_w = min(40, max(24, usable * 0.16))
    col_w = (usable - label_w) / max(n_cols, 1)
    line_h = 6
    cell_h = 12

    # Header row
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(249, 250, 251)
    pdf.set_text_color(*INK_MUTED)
    pdf.set_draw_color(*HAIRLINE)
    pdf.cell(label_w, line_h, "", border=1, fill=True)
    for h in col_headers:
        pdf.cell(col_w, line_h, _truncate(pdf, h, col_w - 2), border=1, align="C", fill=True)
    pdf.ln(line_h)

    # Body rows
    for r, rh in enumerate(row_headers):
        x0 = pdf.get_x()
        y0 = pdf.get_y()
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*INK)
        pdf.set_fill_color(255, 255, 255)
        pdf.cell(label_w, cell_h, _truncate(pdf, rh, label_w - 2), border=1)
        pdf.set_font("Helvetica", "", 8)
        for c in range(len(col_headers)):
            entry = lookup(r, c)
            x = pdf.get_x()
            y = pdf.get_y()
            if entry:
                rr, gg, bb = _hex_to_rgb(entry["shift"].get("color"))
                pdf.set_fill_color(rr, gg, bb)
                pdf.set_text_color(*INK)
                pdf.rect(x, y, col_w, cell_h)
                pdf.set_xy(x + 1, y + 1.5)
                pdf.multi_cell(
                    col_w - 2,
                    3.6,
                    _pdf_safe(entry["text"]),
                    border=0,
                    align="C",
                    fill=True,
                    max_line_height=3.6,
                )
                pdf.set_xy(x + col_w, y)
            elif is_leave(r, c):
                pdf.set_draw_color(*HAIRLINE)
                pdf.set_dash_pattern(dash=1.2, gap=1)
                pdf.rect(x, y, col_w, cell_h)
                pdf.set_dash_pattern()
                pdf.set_font("Helvetica", "I", 7)
                pdf.set_text_color(*INK_MUTED)
                pdf.set_xy(x, y + cell_h / 2 - 2)
                pdf.cell(col_w, 4, "Leave", align="C")
                pdf.set_font("Helvetica", "", 8)
                pdf.set_xy(x + col_w, y)
            else:
                pdf.cell(col_w, cell_h, "", border=1)
        pdf.set_xy(x0, y0 + cell_h)


def _truncate(pdf: FPDF, text: str, max_w: float) -> str:
    text = _pdf_safe(text)
    if pdf.get_string_width(text) <= max_w:
        return text
    while text and pdf.get_string_width(text + "…") > max_w:
        text = text[:-1]
    return text + "…"


def _hex_to_rgb(hex_color: str | None):
    if not hex_color:
        return (243, 244, 246)
    h = hex_color.lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    try:
        r = int(h[0:2], 16)
        g = int(h[2:4], 16)
        b = int(h[4:6], 16)
    except ValueError:
        return (243, 244, 246)
    # Lighten toward white so text stays legible on the fill (~15% tint).
    tint = 0.82
    return (
        round(r + (255 - r) * tint),
        round(g + (255 - g) * tint),
        round(b + (255 - b) * tint),
    )


# --- Excel -----------------------------------------------------------------


def build_rota_xlsx(
    *,
    venue_name: str,
    week_start: date,
    shifts: list[dict],
    staff: list[dict],
    assignments: list[dict],
    leave: dict[str, set[int]] | None = None,
    orientation: str = "staff-rows",
    shift_days: dict | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    leave = leave or {}
    cell = _build_matrix(shifts, assignments, shift_days)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rota"

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F9FAFB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Title block
    ws["A1"] = venue_name
    ws["A1"].font = Font(bold=True, size=14, color="111827")
    ws["A2"] = f"Rota for week {_week_range_label(week_start)}"
    ws["A2"].font = Font(size=11, color="6B7280")

    start_row = 4

    if orientation == "day-rows":
        col_headers = [_name_with_tag(s) for s in staff]
        col_keys = [s["id"] for s in staff]
        # No natural role axis once days are the rows — one flat group.
        row_groups = [(None, [(DAY_NAMES[i], i) for i in range(7)])]

        def lookup(r: int, c: int):
            return cell.get((col_keys[c], r))

        def is_leave(r: int, c: int) -> bool:
            return r in leave.get(col_keys[c], ())
    else:
        groups: dict[str, list[dict]] = {}
        for s in staff:
            groups.setdefault(s.get("role") or "Staff", []).append(s)
        row_groups = [
            (role, [(_name_with_tag(m), m["id"]) for m in members]) for role, members in groups.items()
        ]
        col_headers = [DAY_NAMES[i] for i in range(7)]

        def lookup(r_key: str, c: int):
            return cell.get((r_key, c))

        def is_leave(r_key: str, c: int) -> bool:
            return c in leave.get(r_key, ())

    # Header row
    ws.cell(row=start_row, column=1, value="").border = border
    for c, h in enumerate(col_headers):
        cellref = ws.cell(row=start_row, column=c + 2, value=h)
        cellref.font = Font(bold=True, size=10, color="6B7280")
        cellref.fill = header_fill
        cellref.border = border
        cellref.alignment = center

    # Body, grouped by role (day-rows has a single unlabelled group)
    row = start_row
    for role, rows in row_groups:
        if role is not None:
            row += 1
            role_cell = ws.cell(row=row, column=1, value=role)
            role_cell.font = Font(bold=True, size=10, color="6B7280")
        for label_text, row_key in rows:
            row += 1
            label = ws.cell(row=row, column=1, value=label_text)
            label.font = Font(bold=True, size=10, color="111827")
            label.border = border
            label.alignment = center
            for c in range(len(col_headers)):
                # day-rows passes the day index as row_key via lookup(r, c);
                # staff-rows passes the staff id — both lookup/is_leave accept
                # whichever key this orientation produced.
                entry = lookup(row_key, c)
                on_leave = is_leave(row_key, c)
                text = entry["text"] if entry else ""
                cref = ws.cell(row=row, column=c + 2, value=text or ("Leave" if on_leave and not entry else ""))
                cref.border = border
                cref.alignment = center
                if entry:
                    cref.font = Font(size=9, color="111827")
                    fill_hex = _tint_hex(entry["shift"].get("color"))
                    cref.fill = PatternFill("solid", fgColor=fill_hex)
                elif on_leave:
                    cref.font = Font(size=9, italic=True, color="9CA3AF")
                else:
                    cref.font = Font(size=9, color="111827")

    ws.column_dimensions["A"].width = 24
    for c in range(len(col_headers)):
        ws.column_dimensions[get_column_letter(c + 2)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _tint_hex(hex_color: str | None) -> str:
    r, g, b = _hex_to_rgb(hex_color)
    return f"{r:02X}{g:02X}{b:02X}"
